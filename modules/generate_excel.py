import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook

def create_excel(csv_file,excel_sheet):
    # Update the excel file for report data
    report_csv_data = pd.read_csv(csv_file)
    report_csv_writer = pd.ExcelWriter('reports.xlsx', engine='openpyxl', mode='a', if_sheet_exists='overlay')
    report_csv_data.to_excel(report_csv_writer, sheet_name=excel_sheet, index=False, header=True, startrow=0)
    report_csv_writer.save()

    # Create Filters
    report_workbook = load_workbook(filename = 'reports.xlsx')
    report_worksheet = report_workbook[excel_sheet]
    report_worksheet.auto_filter.ref = report_worksheet.dimensions
    report_workbook.save('reports.xlsx')
    report_workbook.close()